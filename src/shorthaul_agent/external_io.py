"""External data contracts and CSV adapters for operational use."""

from __future__ import annotations

import csv
import json
import re
from datetime import datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd

from shorthaul_agent.models import Instance
from shorthaul_agent.time_utils import parse_time_to_minutes


CSV_SCHEMAS: dict[str, list[dict[str, str]]] = {
    "fleets.csv": [
        {"name": "id", "required": "yes", "description": "Fleet identifier used by routes."},
        {"name": "vehicle_count", "required": "yes", "description": "Number of owned vehicles in this fleet."},
        {"name": "fixed_cost", "required": "no", "description": "Daily or planning-horizon fixed cost."},
        {"name": "variable_cost_per_trip", "required": "no", "description": "Owned-vehicle cost per assigned trip."},
        {"name": "normal_load_minutes", "required": "no", "description": "Loading time without container."},
        {"name": "normal_unload_minutes", "required": "no", "description": "Unloading time without container."},
        {"name": "container_load_minutes", "required": "no", "description": "Loading time with container."},
        {"name": "container_unload_minutes", "required": "no", "description": "Unloading time with container."},
    ],
    "routes.csv": [
        {"name": "id", "required": "yes", "description": "Unique route id."},
        {"name": "origin", "required": "yes", "description": "Origin site or depot."},
        {"name": "destination", "required": "yes", "description": "Destination station or customer cluster."},
        {"name": "wave", "required": "yes", "description": "Dispatch wave label, such as 0600 or AM-1."},
        {"name": "latest_dispatch_minute", "required": "yes", "description": "Minute offset from planning start; 1800 means D+1 06:00."},
        {"name": "travel_minutes", "required": "yes", "description": "Line-haul travel duration."},
        {"name": "fleet_id", "required": "yes", "description": "Fleet id that can serve this route."},
        {"name": "variable_cost", "required": "no", "description": "Route-level owned-trip cost."},
        {"name": "external_cost", "required": "no", "description": "Explicit external-carrier cost. Optional."},
        {"name": "external_cost_multiplier", "required": "no", "description": "Multiplier when external_cost is omitted."},
    ],
    "forecast.csv": [
        {"name": "route_id", "required": "yes", "description": "Route id matching routes.csv."},
        {"name": "minute", "required": "yes", "description": "Demand time bucket as minute offset; 1380 means D+0 23:00."},
        {"name": "volume", "required": "yes", "description": "Forecast demand volume for the bucket."},
    ],
    "milk_run_pairs.csv": [
        {"name": "left_destination", "required": "yes", "description": "First destination in a compatible pair."},
        {"name": "right_destination", "required": "yes", "description": "Second destination in a compatible pair."},
    ],
}


CSV_TEMPLATES: dict[str, str] = {
    "fleets.csv": (
        "id,vehicle_count,fixed_cost,variable_cost_per_trip,normal_load_minutes,normal_unload_minutes,"
        "container_load_minutes,container_unload_minutes\n"
        "Fleet-A,3,520,80,45,45,20,20\n"
        "Fleet-B,2,480,70,45,45,20,20\n"
    ),
    "routes.csv": (
        "id,origin,destination,wave,latest_dispatch_minute,travel_minutes,fleet_id,variable_cost,external_cost,external_cost_multiplier\n"
        "Site-A - Stop-01 - 0600,Site-A,Stop-01,0600,1800,35,Fleet-A,150,,1.35\n"
        "Site-A - Stop-02 - 0600,Site-A,Stop-02,0600,1800,25,Fleet-A,120,,1.35\n"
        "Site-B - Stop-03 - 1400,Site-B,Stop-03,1400,2280,28,Fleet-B,110,,1.35\n"
    ),
    "forecast.csv": (
        "route_id,minute,volume\n"
        "Site-A - Stop-01 - 0600,1380,600\n"
        "Site-A - Stop-01 - 0600,1460,500\n"
        "Site-A - Stop-02 - 0600,1430,400\n"
        "Site-B - Stop-03 - 1400,2100,900\n"
    ),
    "milk_run_pairs.csv": "left_destination,right_destination\nStop-01,Stop-02\n",
}

WORKBOOK_SHEETS: dict[str, dict[str, Any]] = {
    "fleets": {
        "title": "fleets / 车队",
        "required": True,
        "description": "One row per owned fleet.",
        "columns": [
            ("fleet_id", "yes", "Fleet id referenced by routes."),
            ("owned_vehicles", "yes", "Number of owned vehicles."),
            ("fixed_cost", "no", "Daily or planning-horizon fixed cost."),
            ("trip_cost", "no", "Owned-vehicle cost per assigned trip."),
            ("load_min", "no", "Loading minutes without container."),
            ("unload_min", "no", "Unloading minutes without container."),
            ("container_load_min", "no", "Loading minutes with container."),
            ("container_unload_min", "no", "Unloading minutes with container."),
        ],
        "rows": [
            ["Fleet-A", 3, 520, 80, 45, 45, 20, 20],
            ["Fleet-B", 2, 480, 70, 45, 45, 20, 20],
        ],
    },
    "routes": {
        "title": "routes / 线路",
        "required": True,
        "description": "One row per dispatch lane and wave.",
        "columns": [
            ("route_id", "yes", "Unique route id."),
            ("origin", "yes", "Origin site or depot."),
            ("destination", "yes", "Destination station or customer cluster."),
            ("wave", "yes", "Wave label, such as 0600, 1400, AM-1."),
            ("latest_dispatch_time", "yes", "Minute offset or D+n HH:MM. Example: D+1 06:00."),
            ("travel_min", "yes", "Line-haul travel minutes."),
            ("fleet_id", "yes", "Fleet id that can serve this route."),
            ("owned_trip_cost", "no", "Route-level owned-trip cost."),
            ("external_trip_cost", "no", "Explicit external-carrier cost."),
            ("external_multiplier", "no", "Multiplier when external_trip_cost is omitted."),
        ],
        "rows": [
            ["Site-A - Stop-01 - 0600", "Site-A", "Stop-01", "0600", "D+1 06:00", 35, "Fleet-A", 150, "", 1.35],
            ["Site-A - Stop-02 - 0600", "Site-A", "Stop-02", "0600", "D+1 06:00", 25, "Fleet-A", 120, "", 1.35],
            ["Site-B - Stop-03 - 1400", "Site-B", "Stop-03", "1400", "D+1 14:00", 28, "Fleet-B", 110, "", 1.35],
        ],
    },
    "demand": {
        "title": "demand / 货量",
        "required": True,
        "description": "Demand volume by route. ready_time can be omitted; the adapter then uses latest_dispatch_time - 120 minutes.",
        "columns": [
            ("route_id", "yes", "Route id from routes."),
            ("volume", "yes", "Forecast demand volume."),
            ("ready_time", "no", "Minute offset or D+n HH:MM when the volume is available."),
        ],
        "rows": [
            ["Site-A - Stop-01 - 0600", 600, "D+0 23:00"],
            ["Site-A - Stop-01 - 0600", 500, "D+1 00:20"],
            ["Site-A - Stop-02 - 0600", 400, "D+1 00:10"],
            ["Site-B - Stop-03 - 1400", 900, "D+1 11:00"],
        ],
    },
    "compatibility": {
        "title": "compatibility / 串点兼容",
        "required": False,
        "description": "Optional destination pairs that can be served together.",
        "columns": [
            ("left_destination", "yes", "First compatible destination."),
            ("right_destination", "yes", "Second compatible destination."),
        ],
        "rows": [["Stop-01", "Stop-02"]],
    },
    "settings": {
        "title": "settings / 参数",
        "required": False,
        "description": "Optional optimization settings. Missing values use project defaults.",
        "columns": [
            ("key", "yes", "Setting key."),
            ("value", "yes", "Setting value."),
        ],
        "rows": [
            ["vehicle_capacity", 1000],
            ["container_capacity", 800],
            ["max_stops", 3],
            ["allow_container", True],
            ["allow_external", True],
            ["tail_cover_strategy", "cost_aware"],
            ["cost_weight", 1.0],
            ["turnover_weight", 0.5],
            ["fill_rate_weight", 0.2],
        ],
    },
}

SHEET_ALIASES = {
    "fleets": {"fleets", "fleet", "车队", "车辆", "自有车"},
    "routes": {"routes", "route", "线路", "线路表", "lanes", "lanes/routes"},
    "demand": {"demand", "forecast", "货量", "需求", "预测", "volume", "volumes"},
    "compatibility": {"compatibility", "milk_run_pairs", "串点", "串点兼容", "可串点", "pairs"},
    "settings": {"settings", "config", "配置", "参数", "约束", "目标"},
}

FLEET_COLUMN_ALIASES = {
    "id": {"id", "fleet_id", "fleet", "车队编码", "车队id", "车队"},
    "vehicle_count": {"vehicle_count", "owned_vehicles", "owned_vehicle_count", "vehicles", "自有车数量", "车辆数"},
    "fixed_cost": {"fixed_cost", "固定成本"},
    "variable_cost_per_trip": {"variable_cost_per_trip", "trip_cost", "owned_trip_cost", "单趟成本", "自有车单趟成本"},
    "normal_load_minutes": {"normal_load_minutes", "load_min", "loading_min", "装车分钟", "装车时间"},
    "normal_unload_minutes": {"normal_unload_minutes", "unload_min", "unloading_min", "卸车分钟", "卸车时间"},
    "container_load_minutes": {"container_load_minutes", "container_load_min", "容器装车分钟"},
    "container_unload_minutes": {"container_unload_minutes", "container_unload_min", "容器卸车分钟"},
}

ROUTE_COLUMN_ALIASES = {
    "id": {"id", "route_id", "route", "线路编码", "线路id", "线路"},
    "origin": {"origin", "from", "起点", "始发地", "场地", "始发场地"},
    "destination": {"destination", "to", "终点", "目的地", "站点", "目的站点"},
    "wave": {"wave", "波次", "班次", "发运波次"},
    "latest_dispatch_minute": {"latest_dispatch_minute", "latest_dispatch_time", "deadline", "最晚发运时间", "发运截止"},
    "travel_minutes": {"travel_minutes", "travel_min", "duration_min", "在途分钟", "行驶分钟", "在途时长"},
    "fleet_id": {"fleet_id", "fleet", "车队编码", "车队id", "车队"},
    "variable_cost": {"variable_cost", "owned_trip_cost", "trip_cost", "自有变动成本", "自有车成本"},
    "external_cost": {"external_cost", "external_trip_cost", "外部承运商成本", "外部成本"},
    "external_cost_multiplier": {"external_cost_multiplier", "external_multiplier", "外部成本倍率"},
}

DEMAND_COLUMN_ALIASES = {
    "route_id": {"route_id", "route", "线路编码", "线路id", "线路"},
    "minute": {"minute", "ready_time", "demand_time", "arrival_time", "货量产生时间", "可发运时间", "时间"},
    "volume": {"volume", "forecast_volume", "demand", "packages", "包裹量", "货量", "需求量"},
}

PAIR_COLUMN_ALIASES = {
    "left_destination": {"left_destination", "left", "destination_a", "站点1", "站点编码1", "左站点"},
    "right_destination": {"right_destination", "right", "destination_b", "站点2", "站点编码2", "右站点"},
}


def schema_payload() -> dict[str, Any]:
    return {
        "format": "shorthaul-agent/v1",
        "recommended_input": "single workbook: shorthaul_dispatch_template.xlsx",
        "required_workbook_sheets": ["fleets", "routes", "demand"],
        "optional_workbook_sheets": ["compatibility", "settings"],
        "required_files": ["fleets.csv", "routes.csv", "forecast.csv"],
        "optional_files": ["milk_run_pairs.csv", "config_overrides.json"],
        "workbook_schemas": WORKBOOK_SHEETS,
        "csv_schemas": CSV_SCHEMAS,
        "time_fields": {
            "minute_offset": "Integer minutes from planning start. Example: 1800 means D+1 06:00.",
            "clock_text": "HH:MM and HHMM are accepted in JSON; CSV templates use minute offsets to avoid day-boundary ambiguity.",
        },
        "api_payload": {
            "request": "Natural-language dispatch requirement.",
            "prefer_cpsat": "Boolean. True uses CP-SAT when available; false forces heuristic.",
            "config_overrides": "ProblemConfig overrides such as capacity and objective weights.",
            "instance": "Scenario object containing fleets, routes, and forecast buckets.",
        },
        "endpoints": {
            "json_solve": "POST /schedule",
            "multipart_upload_solve": "POST /schedule/upload",
            "human_contract": "GET /contract",
            "template_preview": "GET /templates/view",
            "workbook_template": "GET /templates/workbook.xlsx",
            "server_local_csv_solve": "POST /schedule/from-csv-dir",
            "validate_json": "POST /validate-instance",
        },
    }


def build_payload_from_workbook(
    workbook_path: str | Path,
    request_text: str,
    *,
    instance_id: str = "external-workbook-instance",
    date: str = "",
    prefer_cpsat: bool = True,
    config_overrides: dict[str, Any] | None = None,
) -> dict[str, Any]:
    instance = load_instance_from_workbook(workbook_path, instance_id=instance_id, date=date)
    overrides = load_config_overrides_from_workbook(workbook_path)
    if config_overrides:
        overrides = _deep_merge(overrides, config_overrides)
    return {
        "request": request_text,
        "prefer_cpsat": prefer_cpsat,
        "config_overrides": overrides,
        "instance": instance.to_dict(),
    }


def load_instance_from_workbook(workbook_path: str | Path, *, instance_id: str = "external-workbook-instance", date: str = "") -> Instance:
    sheets = _read_workbook_sheets(Path(workbook_path))
    missing = [sheet for sheet in ("fleets", "routes", "demand") if sheet not in sheets]
    if missing:
        raise ValueError(f"Missing required workbook sheets: {', '.join(missing)}")

    fleets = [_normalize_workbook_fleet(row) for row in _frame_rows(sheets["fleets"], FLEET_COLUMN_ALIASES)]
    routes = [_normalize_workbook_route(row) for row in _frame_rows(sheets["routes"], ROUTE_COLUMN_ALIASES)]
    routes_by_id = {route["id"]: route for route in routes}
    forecast = [_normalize_workbook_forecast(row, routes_by_id) for row in _frame_rows(sheets["demand"], DEMAND_COLUMN_ALIASES)]
    return Instance.from_dict({"id": instance_id, "date": date, "fleets": fleets, "routes": routes, "forecast": forecast})


def load_config_overrides_from_workbook(workbook_path: str | Path) -> dict[str, Any]:
    sheets = _read_workbook_sheets(Path(workbook_path))
    overrides: dict[str, Any] = {}

    if "settings" in sheets:
        for row in _frame_rows(sheets["settings"], {"key": {"key", "参数", "字段"}, "value": {"value", "值", "配置值"}}):
            key = _clean(row.get("key"))
            if not key:
                continue
            _assign_setting(overrides, key, _parse_setting_value(row.get("value")))

    if "compatibility" in sheets:
        pairs = []
        for row in _frame_rows(sheets["compatibility"], PAIR_COLUMN_ALIASES):
            left = _clean(row.get("left_destination"))
            right = _clean(row.get("right_destination"))
            if left and right:
                pairs.append([left, right])
        if pairs:
            overrides["milk_run_pairs"] = pairs
    return overrides


def workbook_template_bytes() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["ShortHaul Dispatch Agent 输入模板"])
    ws.append(["使用步骤"])
    ws.append(["1. 填 fleets、routes、demand 三张必需表。"])
    ws.append(["2. compatibility 和 settings 可按需填写。"])
    ws.append(["3. 在 Web UI 上传本文件并点击运行。"])
    ws.append(["时间字段支持分钟偏移或 D+n HH:MM，例如 D+1 06:00。"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 72

    header_fill = PatternFill("solid", fgColor="EAF4F4")
    for sheet_name, meta in WORKBOOK_SHEETS.items():
        sheet = wb.create_sheet(sheet_name)
        headers = [column[0] for column in meta["columns"]]
        sheet.append(headers)
        for row in meta["rows"]:
            sheet.append(row)
        for idx, column in enumerate(headers, start=1):
            cell = sheet.cell(1, idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[get_column_letter(idx)].width = max(14, min(28, len(column) + 6))
        sheet.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def write_workbook_template(path: str | Path) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(workbook_template_bytes())


def write_workbook_from_payload(payload: dict[str, Any], path: str | Path) -> None:
    instance = payload["instance"]
    config = payload.get("config_overrides", {})
    rows = {
        "fleets": [
            [
                fleet.get("id"),
                fleet.get("vehicle_count"),
                fleet.get("fixed_cost"),
                fleet.get("variable_cost_per_trip"),
                fleet.get("normal_load_minutes"),
                fleet.get("normal_unload_minutes"),
                fleet.get("container_load_minutes"),
                fleet.get("container_unload_minutes"),
            ]
            for fleet in instance.get("fleets", [])
        ],
        "routes": [
            [
                route.get("id"),
                route.get("origin"),
                route.get("destination"),
                route.get("wave"),
                route.get("latest_dispatch_minute"),
                route.get("travel_minutes"),
                route.get("fleet_id"),
                route.get("variable_cost"),
                route.get("external_cost", ""),
                route.get("external_cost_multiplier", ""),
            ]
            for route in instance.get("routes", [])
        ],
        "demand": [
            [bucket.get("route_id"), bucket.get("volume"), bucket.get("minute")]
            for bucket in instance.get("forecast", [])
        ],
        "compatibility": [
            [left, right] for left, right in config.get("milk_run_pairs", [])
        ],
        "settings": _settings_rows(config),
    }
    _write_workbook_rows(rows, Path(path))


def _write_workbook_rows(rows: dict[str, list[list[Any]]], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="EAF4F4")
    for sheet_name, meta in WORKBOOK_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        headers = [column[0] for column in meta["columns"]]
        ws.append(headers)
        for row in rows.get(sheet_name, []):
            ws.append(row)
        for idx, column in enumerate(headers, start=1):
            cell = ws.cell(1, idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(idx)].width = max(14, min(28, len(column) + 6))
        ws.freeze_panes = "A2"
    wb.save(path)


def _settings_rows(config: dict[str, Any]) -> list[list[Any]]:
    rows = []
    objective = config.get("objective_weights", {})
    for key, value in config.items():
        if key in {"milk_run_pairs", "objective_weights"}:
            continue
        rows.append([key, value])
    for key, value in objective.items():
        rows.append([f"{key}_weight" if key != "fill_rate" else "fill_rate_weight", value])
    return rows


def render_contract_html() -> str:
    rows = []
    for sheet_name, meta in WORKBOOK_SHEETS.items():
        columns = "".join(
            f"<tr><td><code>{name}</code></td><td>{required}</td><td>{description}</td></tr>"
            for name, required, description in meta["columns"]
        )
        rows.append(
            f"""
            <section>
              <h2>{meta['title']}</h2>
              <p>{meta['description']}</p>
              <table>
                <thead><tr><th>字段</th><th>必填</th><th>说明</th></tr></thead>
                <tbody>{columns}</tbody>
              </table>
            </section>
            """
        )
    return _page_html(
        "ShortHaul 输入契约",
        f"""
        <p class="lead">推荐外部使用者只准备一个 Excel 工作簿：<code>fleets</code>、<code>routes</code>、<code>demand</code>
        三张表即可运行；<code>compatibility</code> 和 <code>settings</code> 用于补充串点关系和优化参数。</p>
        <div class="actions">
          <a href="/templates/workbook.xlsx">下载 Excel 模板</a>
          <a href="/templates/view">查看模板样例</a>
          <a href="/">返回调度页面</a>
        </div>
        <ol>
          <li>从 TMS、Excel 或数据库导出车队、线路、货量三类数据。</li>
          <li>按模板列名粘贴到对应 sheet。列名支持英文模板名，也兼容常见中文列名。</li>
          <li>时间可写分钟偏移，例如 <code>1800</code>，也可写 <code>D+1 06:00</code>。</li>
          <li>上传工作簿并运行，系统会自动转换为内部 JSON、生成任务、调用求解器并展示甘特图。</li>
        </ol>
        <section>
          <h2>常见业务字段映射</h2>
          <table>
            <thead><tr><th>业务数据</th><th>工作表</th><th>模板字段</th></tr></thead>
            <tbody>
              <tr><td>车队编号、自有车数量</td><td><code>fleets</code></td><td><code>fleet_id</code>, <code>owned_vehicles</code></td></tr>
              <tr><td>线路编号、始发地、目的地、波次</td><td><code>routes</code></td><td><code>route_id</code>, <code>origin</code>, <code>destination</code>, <code>wave</code></td></tr>
              <tr><td>最晚发运时间、在途分钟</td><td><code>routes</code></td><td><code>latest_dispatch_time</code>, <code>travel_min</code></td></tr>
              <tr><td>预测货量、包裹量</td><td><code>demand</code></td><td><code>route_id</code>, <code>volume</code>, <code>ready_time</code> 可选</td></tr>
              <tr><td>可串点站点对</td><td><code>compatibility</code></td><td><code>left_destination</code>, <code>right_destination</code></td></tr>
            </tbody>
          </table>
        </section>
        {''.join(rows)}
        """,
    )


def render_templates_html() -> str:
    workbook_sections = []
    for sheet_name, meta in WORKBOOK_SHEETS.items():
        headers = [column[0] for column in meta["columns"]]
        body = "".join(
            "<tr>" + "".join(f"<td>{value}</td>" for value in row) + "</tr>"
            for row in meta["rows"][:4]
        )
        workbook_sections.append(
            f"""
            <section>
              <h2>{sheet_name}.sheet</h2>
              <p>{meta['description']}</p>
              <table>
                <thead><tr>{''.join(f'<th>{header}</th>' for header in headers)}</tr></thead>
                <tbody>{body}</tbody>
              </table>
            </section>
            """
        )
    csv_links = "".join(f'<li><a href="/templates/csv/{name}">{name}</a></li>' for name in CSV_TEMPLATES)
    return _page_html(
        "ShortHaul 模板预览",
        f"""
        <p class="lead">默认推荐下载单个 Excel 工作簿。CSV 模板保留给需要系统集成或自动化导出的场景。</p>
        <div class="actions">
          <a href="/templates/workbook.xlsx">下载 Excel 工作簿模板</a>
          <a href="/contract">查看字段说明</a>
          <a href="/">返回调度页面</a>
        </div>
        {''.join(workbook_sections)}
        <section>
          <h2>高级：CSV 模板</h2>
          <ul>{csv_links}</ul>
        </section>
        """,
    )


def build_payload_from_csv_dir(
    data_dir: str | Path,
    request_text: str,
    *,
    instance_id: str = "external-instance",
    date: str = "",
    prefer_cpsat: bool = True,
    config_overrides: dict[str, Any] | None = None,
) -> dict[str, Any]:
    data_path = Path(data_dir)
    instance = load_instance_from_csv_dir(data_path, instance_id=instance_id, date=date)
    overrides = load_config_overrides_from_csv_dir(data_path)
    if config_overrides:
        overrides.update(config_overrides)
    return {
        "request": request_text,
        "prefer_cpsat": prefer_cpsat,
        "config_overrides": overrides,
        "instance": instance.to_dict(),
    }


def load_instance_from_csv_dir(data_dir: str | Path, *, instance_id: str = "external-instance", date: str = "") -> Instance:
    data_path = Path(data_dir)
    fleets = [_normalize_fleet(row) for row in _read_required_csv(data_path / "fleets.csv")]
    routes = [_normalize_route(row) for row in _read_required_csv(data_path / "routes.csv")]
    forecast = [_normalize_forecast(row) for row in _read_required_csv(data_path / "forecast.csv")]
    return Instance.from_dict({"id": instance_id, "date": date, "fleets": fleets, "routes": routes, "forecast": forecast})


def load_config_overrides_from_csv_dir(data_dir: str | Path) -> dict[str, Any]:
    data_path = Path(data_dir)
    overrides: dict[str, Any] = {}
    config_path = data_path / "config_overrides.json"
    if config_path.exists():
        overrides.update(json.loads(config_path.read_text(encoding="utf-8")))

    pairs_path = data_path / "milk_run_pairs.csv"
    if pairs_path.exists():
        pairs = []
        for row in _read_required_csv(pairs_path):
            left = _clean(row.get("left_destination"))
            right = _clean(row.get("right_destination"))
            if left and right:
                pairs.append([left, right])
        if pairs:
            overrides["milk_run_pairs"] = pairs
    return overrides


def _read_required_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"Missing required CSV file: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp)
        if not reader.fieldnames:
            raise ValueError(f"CSV file has no header: {path}")
        return [dict(row) for row in reader]


def _normalize_fleet(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": _required(row, "id"),
        "vehicle_count": _int(row, "vehicle_count"),
        "fixed_cost": _int(row, "fixed_cost", 600),
        "variable_cost_per_trip": _int(row, "variable_cost_per_trip", 80),
        "normal_load_minutes": _int(row, "normal_load_minutes", 45),
        "normal_unload_minutes": _int(row, "normal_unload_minutes", 45),
        "container_load_minutes": _int(row, "container_load_minutes", 20),
        "container_unload_minutes": _int(row, "container_unload_minutes", 20),
    }


def _normalize_route(row: dict[str, Any]) -> dict[str, Any]:
    payload = {
        "id": _required(row, "id"),
        "origin": _required(row, "origin"),
        "destination": _required(row, "destination"),
        "wave": _required(row, "wave"),
        "latest_dispatch_minute": _minute(row, "latest_dispatch_minute"),
        "travel_minutes": _int(row, "travel_minutes"),
        "fleet_id": _required(row, "fleet_id"),
        "variable_cost": _int(row, "variable_cost", 120),
        "external_cost_multiplier": _float(row, "external_cost_multiplier", 1.35),
    }
    external_cost = _optional_int(row, "external_cost")
    if external_cost is not None:
        payload["external_cost"] = external_cost
    return payload


def _normalize_forecast(row: dict[str, Any]) -> dict[str, Any]:
    return {"route_id": _required(row, "route_id"), "minute": _minute(row, "minute"), "volume": _int(row, "volume")}


def _required(row: dict[str, Any], key: str) -> str:
    value = _clean(row.get(key))
    if not value:
        raise ValueError(f"Missing required field: {key}")
    return value


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _int(row: dict[str, Any], key: str, default: int | None = None) -> int:
    value = _clean(row.get(key))
    if not value:
        if default is None:
            raise ValueError(f"Missing required integer field: {key}")
        return default
    return int(round(float(value)))


def _optional_int(row: dict[str, Any], key: str) -> int | None:
    value = _clean(row.get(key))
    return None if not value else int(round(float(value)))


def _float(row: dict[str, Any], key: str, default: float) -> float:
    value = _clean(row.get(key))
    return default if not value else float(value)


def _minute(row: dict[str, Any], key: str) -> int:
    value = _clean(row.get(key))
    if not value:
        raise ValueError(f"Missing required time field: {key}")
    if value.lstrip("-").isdigit():
        return int(value)
    return parse_time_to_minutes(value)


def _read_workbook_sheets(path: Path) -> dict[str, pd.DataFrame]:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    raw_sheets = pd.read_excel(path, sheet_name=None)
    normalized: dict[str, pd.DataFrame] = {}
    for sheet_name, frame in raw_sheets.items():
        key = _canonical_sheet_name(sheet_name)
        if key:
            normalized[key] = frame.dropna(how="all")
    return normalized


def _canonical_sheet_name(sheet_name: str) -> str:
    cleaned = str(sheet_name).strip().lower()
    for canonical, aliases in SHEET_ALIASES.items():
        if cleaned in {alias.lower() for alias in aliases}:
            return canonical
    return ""


def _frame_rows(frame: pd.DataFrame, aliases: dict[str, set[str]]) -> list[dict[str, Any]]:
    if frame.empty:
        return []
    renamed = {}
    alias_lookup = {
        _normalize_column(alias): canonical
        for canonical, names in aliases.items()
        for alias in names
    }
    for column in frame.columns:
        key = alias_lookup.get(_normalize_column(column))
        if key:
            renamed[column] = key
    normalized = frame.rename(columns=renamed)
    rows = []
    for raw in normalized.to_dict(orient="records"):
        row = {key: _empty_if_nan(value) for key, value in raw.items()}
        if any(_clean(value) for value in row.values()):
            rows.append(row)
    return rows


def _normalize_column(value: Any) -> str:
    return str(value).strip().lower().replace(" ", "_")


def _empty_if_nan(value: Any) -> Any:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass
    return value


def _normalize_workbook_fleet(row: dict[str, Any]) -> dict[str, Any]:
    return _normalize_fleet(
        {
            "id": row.get("id"),
            "vehicle_count": row.get("vehicle_count"),
            "fixed_cost": row.get("fixed_cost"),
            "variable_cost_per_trip": row.get("variable_cost_per_trip"),
            "normal_load_minutes": row.get("normal_load_minutes"),
            "normal_unload_minutes": row.get("normal_unload_minutes"),
            "container_load_minutes": row.get("container_load_minutes"),
            "container_unload_minutes": row.get("container_unload_minutes"),
        }
    )


def _normalize_workbook_route(row: dict[str, Any]) -> dict[str, Any]:
    payload = {
        "id": _required(row, "id"),
        "origin": _required(row, "origin"),
        "destination": _required(row, "destination"),
        "wave": _required(row, "wave"),
        "latest_dispatch_minute": _operational_minute(row.get("latest_dispatch_minute")),
        "travel_minutes": _int(row, "travel_minutes"),
        "fleet_id": _required(row, "fleet_id"),
        "variable_cost": _int(row, "variable_cost", 120),
        "external_cost_multiplier": _float(row, "external_cost_multiplier", 1.35),
    }
    external_cost = _optional_int(row, "external_cost")
    if external_cost is not None:
        payload["external_cost"] = external_cost
    return payload


def _normalize_workbook_forecast(row: dict[str, Any], routes_by_id: dict[str, dict[str, Any]]) -> dict[str, Any]:
    route_id = _required(row, "route_id")
    if route_id not in routes_by_id:
        raise ValueError(f"Demand references unknown route_id: {route_id}")
    minute_value = _clean(row.get("minute"))
    if minute_value:
        minute = _operational_minute(row.get("minute"))
    else:
        minute = max(int(routes_by_id[route_id]["latest_dispatch_minute"]) - 120, 0)
    return {"route_id": route_id, "minute": minute, "volume": _int(row, "volume")}


def _operational_minute(value: Any) -> int:
    value = _empty_if_nan(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(round(float(value)))
    if isinstance(value, pd.Timestamp):
        return value.hour * 60 + value.minute
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    if isinstance(value, timedelta):
        return int(value.total_seconds() // 60)
    text = _clean(value)
    if not text:
        raise ValueError("Missing time value")
    day_match = re.search(r"d\s*\+\s*(\d+)\s*(\d{1,2}):(\d{2})", text, flags=re.IGNORECASE)
    if not day_match:
        day_match = re.search(r"(\d+)\s*d\s+(\d{1,2}):(\d{2})", text, flags=re.IGNORECASE)
    if day_match:
        day = int(day_match.group(1))
        hour = int(day_match.group(2))
        minute = int(day_match.group(3))
        return day * 1440 + hour * 60 + minute
    return parse_time_to_minutes(text)


def _assign_setting(overrides: dict[str, Any], raw_key: str, value: Any) -> None:
    key = _normalize_column(raw_key)
    mapping = {
        "cost_weight": ("objective_weights", "cost"),
        "turnover_weight": ("objective_weights", "turnover"),
        "fill_rate_weight": ("objective_weights", "fill_rate"),
    }
    if key in mapping:
        parent, child = mapping[key]
        overrides.setdefault(parent, {})[child] = value
    else:
        overrides[key] = value


def _parse_setting_value(value: Any) -> Any:
    value = _empty_if_nan(value)
    if isinstance(value, str):
        text = value.strip()
        if text.lower() in {"true", "yes", "1", "是"}:
            return True
        if text.lower() in {"false", "no", "0", "否"}:
            return False
        if text.startswith("{") or text.startswith("["):
            return json.loads(text)
        try:
            return float(text) if "." in text else int(text)
        except ValueError:
            return text
    return value


def _deep_merge(base: dict[str, Any], overrides: dict[str, Any]) -> dict[str, Any]:
    merged = dict(base or {})
    for key, value in (overrides or {}).items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _deep_merge(merged[key], value)
        else:
            merged[key] = value
    return merged


def _page_html(title: str, body: str) -> str:
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{title}</title>
  <style>
    body {{ margin:0; background:#f6f7f9; color:#172033; font:14px/1.6 -apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif; }}
    header {{ background:#111827; color:#fff; padding:20px 32px; }}
    main {{ max-width:1120px; margin:0 auto; padding:24px; }}
    section {{ background:#fff; border:1px solid #d7dce3; border-radius:8px; margin:16px 0; padding:18px; }}
    h1 {{ margin:0; font-size:22px; }}
    h2 {{ margin:0 0 8px; font-size:17px; }}
    .lead {{ background:#eefaf6; border:1px solid #b6e6d6; border-radius:8px; padding:14px; }}
    .actions {{ display:flex; gap:10px; flex-wrap:wrap; margin:16px 0; }}
    .actions a {{ background:#0f766e; color:#fff; padding:8px 12px; border-radius:6px; text-decoration:none; font-weight:650; }}
    table {{ border-collapse:collapse; width:100%; margin-top:10px; }}
    th,td {{ border:1px solid #d7dce3; padding:8px; text-align:left; vertical-align:top; }}
    th {{ background:#fafbfc; }}
    code {{ background:#f1f5f9; padding:1px 5px; border-radius:4px; }}
  </style>
</head>
<body>
  <header><h1>{title}</h1></header>
  <main>{body}</main>
</body>
</html>"""
